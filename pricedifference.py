from openpyxl import load_workbook

def deger_farki(path, sheet_name, start_col=3):
    price_value = load_workbook(filename=path)
    if sheet_name not in price_value.sheetnames:
        print(f"'{sheet_name}' bulunamadı")
        return

    sheet = price_value[sheet_name]

    for row in range(2, sheet.max_row + 1):
        ikinci_sütun_degeri = sheet.cell(row=row, column=2).value
        üçüncü_sütun_degeri = sheet.cell(row=row, column=start_col).value
        dördüncü_sutün_degeri = sheet.cell(row=row, column=start_col + 1).value

        if ikinci_sütun_degeri is not None and üçüncü_sütun_degeri is not None and dördüncü_sutün_degeri is not None:
            result = dördüncü_sutün_degeri - üçüncü_sütun_degeri
            result = round(result, 2)
            print(f"{ikinci_sütun_degeri} = {result}")

            sheet.cell(row=row, column=start_col + 2, value=result)
            
    price_value.save("new_makuldegerfarktablosu.xlsx")

if __name__ == "__main__":
    deger_farki("makuldegertablosu.xlsx", sheet_name="price_value", start_col=3)
